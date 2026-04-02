"""SDLC Checklist parser module for the RAG pipeline.

Parses Software Development Life Cycle checklist documents used across
SDLC phases: requirements, design review, code review, testing,
deployment, and UAT. Supports Markdown, Excel, CSV, TXT, PDF, and DOCX
formats.

Each checklist item becomes a separate chunk with structured metadata
including phase, status (checked/unchecked/na), priority, and assignee.
For Markdown, detects GitHub-style checkboxes (- [ ] / - [x]). For
Excel/CSV, expects columns: item, status, phase, priority, assignee.
For PDF/DOCX, uses bullet and heading heuristics to extract checklist
structure.

The classify_phase() async function runs as a post-parse step in
task_executor.py to auto-detect the SDLC phase when not explicitly
provided.
"""

import logging
import re
from copy import deepcopy
from io import BytesIO

from deepdoc.parser.utils import get_text
from rag.nlp import rag_tokenizer
from common.token_utils import num_tokens_from_string

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# SDLC phase definitions with heading/keyword detection patterns.
# Each phase maps to a regex that matches common section headings.
# ---------------------------------------------------------------------------
SDLC_PHASES = {
    "requirements": re.compile(
        r"(requirement|functional\s+spec|user\s+stor|acceptance\s+criteria|business\s+rule|use\s+case|feature\s+request)",
        re.IGNORECASE,
    ),
    "design_review": re.compile(
        r"(design\s+review|architecture|system\s+design|technical\s+design|hld|lld|component\s+design|api\s+design|data\s+model)",
        re.IGNORECASE,
    ),
    "code_review": re.compile(
        r"(code\s+review|peer\s+review|pull\s+request|merge\s+request|coding\s+standard|static\s+analysis|code\s+quality)",
        re.IGNORECASE,
    ),
    "testing": re.compile(
        r"(test\s+case|test\s+plan|unit\s+test|integration\s+test|regression|smoke\s+test|e2e\s+test|qa\s+checklist|quality\s+assurance|bug\s+verification)",
        re.IGNORECASE,
    ),
    "security": re.compile(
        r"(security|vulnerability|penetration|owasp|threat\s+model|access\s+control|authentication|authorization|encryption|compliance\s+check)",
        re.IGNORECASE,
    ),
    "deployment": re.compile(
        r"(deploy|release|rollback|infrastructure|ci[\s/]*cd|pipeline|staging|production\s+readiness|go[\-\s]*live|migration\s+plan)",
        re.IGNORECASE,
    ),
    "uat": re.compile(
        r"(uat|user\s+acceptance|sign[\-\s]*off|stakeholder\s+review|demo|client\s+validation|end[\-\s]*user\s+test)",
        re.IGNORECASE,
    ),
    "maintenance": re.compile(
        r"(maintenance|post[\-\s]*deploy|monitoring|runbook|incident|support\s+handoff|documentation\s+update|knowledge\s+transfer)",
        re.IGNORECASE,
    ),
}

# Checkbox patterns for Markdown-style checklists
# Matches: - [ ] unchecked, - [x] checked, - [X] checked, - [~] partial/na
MD_CHECKBOX_PATTERN = re.compile(
    r"^[\s]*[-*+]\s*\[([ xX~])\]\s*(.+)$",
    re.MULTILINE,
)

# Priority markers detected inline: (P0), (P1), [HIGH], [CRITICAL], etc.
PRIORITY_PATTERN = re.compile(
    r"\b(?:P([0-4])|\[(critical|high|medium|low|blocker)\])",
    re.IGNORECASE,
)

# Assignee pattern: @username or assigned: Name
ASSIGNEE_PATTERN = re.compile(
    r"(?:@(\w[\w.-]*)|assigned(?:\s*to)?[:\s]+([^,\n\]]+))",
    re.IGNORECASE,
)

# Section heading pattern for Markdown
MD_HEADING_PATTERN = re.compile(r"^(#{1,6})\s+(.+)$", re.MULTILINE)

# Status mapping from checkbox content to normalized status
STATUS_MAP = {
    " ": "unchecked",
    "x": "checked",
    "X": "checked",
    "~": "na",
}

# LLM prompt for phase classification when auto-detection fails
PHASE_CLASSIFICATION_PROMPT = """You are an SDLC phase classifier.
Classify the following checklist document into exactly ONE of these SDLC phases:
- requirements: Requirements gathering, user stories, acceptance criteria
- design_review: Architecture review, system design, API design, data modeling
- code_review: Code review, coding standards, static analysis, pull request checks
- testing: Test plans, test cases, QA checklists, regression testing
- security: Security review, vulnerability assessment, OWASP, threat modeling
- deployment: Release checklists, CI/CD, rollback plans, production readiness
- uat: User acceptance testing, stakeholder sign-off, demo validation
- maintenance: Post-deployment, monitoring, runbooks, support handoff

Respond with ONLY the phase name (one word or underscore-separated, lowercase).

Document title: {title}
Document content (first section):
{content}

Phase:"""

# Maximum tokens sent to LLM for phase classification
MAX_CLASSIFY_TOKENS = 2000

# Target chunk size for paragraph-based fallback splitting
DEFAULT_CHUNK_TOKEN_NUM = 512


def _extract_checkbox_status(marker: str) -> str:
    """Map a checkbox marker character to a normalized status string.

    Args:
        marker: Single character from inside [ ] brackets.

    Returns:
        Normalized status: 'checked', 'unchecked', or 'na'.
    """
    return STATUS_MAP.get(marker, "unchecked")


def _extract_priority(text: str) -> str:
    """Extract priority level from checklist item text.

    Scans for patterns like (P0), (P1), [HIGH], [CRITICAL] inline.
    Maps P-levels and keyword levels to a normalized priority string.

    Args:
        text: Checklist item text content.

    Returns:
        Priority string (critical/high/medium/low) or empty string if not found.
    """
    match = PRIORITY_PATTERN.search(text)
    if not match:
        return ""

    # P-level match (P0-P4)
    if match.group(1) is not None:
        p_level = int(match.group(1))
        return {0: "critical", 1: "high", 2: "medium", 3: "low", 4: "low"}.get(p_level, "medium")

    # Keyword match ([HIGH], [CRITICAL], etc.)
    if match.group(2):
        return match.group(2).strip().lower()

    return ""


def _extract_assignee(text: str) -> str:
    """Extract assignee from checklist item text.

    Looks for @username or 'assigned to: Name' patterns.

    Args:
        text: Checklist item text content.

    Returns:
        Assignee name/username or empty string if not found.
    """
    match = ASSIGNEE_PATTERN.search(text)
    if not match:
        return ""

    # @username match
    if match.group(1):
        return match.group(1).strip()

    # 'assigned to:' match
    if match.group(2):
        return match.group(2).strip()

    return ""


def _detect_phase_from_text(text: str) -> str:
    """Detect the SDLC phase from document text using keyword matching.

    Scans the text against SDLC_PHASES patterns and returns the phase
    with the most keyword hits. Falls back to empty string if no
    phase is detected.

    Args:
        text: Full document text or section heading.

    Returns:
        Phase identifier string or empty string if undetectable.
    """
    scores: dict[str, int] = {}
    for phase, pattern in SDLC_PHASES.items():
        matches = pattern.findall(text)
        if matches:
            scores[phase] = len(matches)

    if not scores:
        return ""

    # Return phase with highest match count
    return max(scores, key=scores.get)


def _detect_phase_from_heading(heading: str) -> str:
    """Detect SDLC phase from a single section heading.

    Args:
        heading: Section heading text.

    Returns:
        Phase identifier string or empty string if undetectable.
    """
    for phase, pattern in SDLC_PHASES.items():
        if pattern.search(heading):
            return phase
    return ""


def _parse_markdown_checklists(text: str) -> list[dict]:
    """Parse Markdown checklist items with GitHub-style checkboxes.

    Splits the document by headings first, then extracts checkbox items
    within each section. Each item gets the section's detected SDLC phase.
    Non-checkbox lines within a section are attached as context to the
    preceding checkbox item.

    Args:
        text: Full Markdown document text.

    Returns:
        List of dicts with keys: item_text, status, phase, priority,
        assignee, section_heading, item_index.
    """
    items = []
    # Split into sections by headings
    sections = _split_md_sections(text)

    item_index = 0
    for heading, content in sections:
        # Detect phase from heading first, fall back to content
        phase = _detect_phase_from_heading(heading) if heading else ""
        if not phase:
            phase = _detect_phase_from_text(content[:500])

        # Extract checkbox items from section content
        lines = content.split("\n")
        pending_context = []

        for line in lines:
            checkbox_match = MD_CHECKBOX_PATTERN.match(line)
            if checkbox_match:
                # Flush pending context to previous item if exists
                if pending_context and items:
                    items[-1]["context"] = "\n".join(pending_context)
                    pending_context = []

                marker = checkbox_match.group(1)
                item_text = checkbox_match.group(2).strip()

                items.append({
                    "item_text": item_text,
                    "status": _extract_checkbox_status(marker),
                    "phase": phase,
                    "priority": _extract_priority(item_text),
                    "assignee": _extract_assignee(item_text),
                    "section_heading": heading,
                    "context": "",
                    "item_index": item_index,
                })
                item_index += 1
            elif line.strip():
                # Non-checkbox line — accumulate as context for next/prev item
                pending_context.append(line.strip())

        # Flush remaining context to last item
        if pending_context and items:
            prev_ctx = items[-1].get("context", "")
            items[-1]["context"] = (prev_ctx + "\n" + "\n".join(pending_context)).strip()

    return items


def _split_md_sections(text: str) -> list[tuple[str, str]]:
    """Split Markdown text into (heading, content) pairs.

    Args:
        text: Full Markdown text.

    Returns:
        List of (heading_text, section_content) tuples.
        First element may have empty heading for preamble content.
    """
    parts = MD_HEADING_PATTERN.split(text)
    sections = []

    # First element is content before any heading (preamble)
    if parts[0].strip():
        sections.append(("", parts[0].strip()))

    # Heading parts come in groups of 3: (hash_marks, heading_text, content)
    i = 1
    while i < len(parts):
        heading_text = parts[i + 1].strip() if i + 1 < len(parts) else ""
        content = parts[i + 2].strip() if i + 2 < len(parts) else ""
        sections.append((heading_text, content))
        i += 3

    return sections


def _parse_tabular_checklist(text: str, delimiter: str = "\t") -> list[dict]:
    """Parse tabular checklist from CSV/TXT with columns.

    Expected columns: item, status, phase, priority, assignee.
    Minimum 2 columns (item + status). Extra columns are optional.

    Args:
        text: Raw text content with delimited columns.
        delimiter: Column separator character.

    Returns:
        List of dicts with checklist item metadata.
    """
    lines = text.strip().split("\n")
    if not lines:
        return []

    items = []
    # Detect header row by checking for common header keywords
    header_keywords = {"item", "task", "checklist", "description", "status", "phase", "priority", "assignee"}
    first_row_lower = lines[0].lower()
    has_header = any(kw in first_row_lower for kw in header_keywords)

    # Parse header to determine column mapping
    col_map = {"item": 0, "status": 1, "phase": 2, "priority": 3, "assignee": 4}
    start_row = 0

    if has_header:
        headers = [h.strip().lower() for h in lines[0].split(delimiter)]
        col_map = {}
        for idx, h in enumerate(headers):
            # Map header names to standard fields
            if any(kw in h for kw in ("item", "task", "checklist", "description", "check")):
                col_map["item"] = idx
            elif "status" in h or "done" in h or "complete" in h:
                col_map["status"] = idx
            elif "phase" in h or "stage" in h:
                col_map["phase"] = idx
            elif "priority" in h or "severity" in h:
                col_map["priority"] = idx
            elif "assign" in h or "owner" in h or "responsible" in h:
                col_map["assignee"] = idx
        # Default item column to 0 if not detected
        if "item" not in col_map:
            col_map["item"] = 0
        start_row = 1

    for row_idx, line in enumerate(lines[start_row:], start=start_row):
        cols = [c.strip() for c in line.split(delimiter)]
        if not cols or not cols[0]:
            continue

        item_col = col_map.get("item", 0)
        item_text = cols[item_col] if item_col < len(cols) else ""
        if not item_text:
            continue

        # Extract status from column or infer from text
        status_col = col_map.get("status")
        status = "unchecked"
        if status_col is not None and status_col < len(cols):
            raw_status = cols[status_col].strip().lower()
            if raw_status in ("done", "yes", "x", "true", "checked", "pass", "passed", "complete", "completed"):
                status = "checked"
            elif raw_status in ("n/a", "na", "skip", "skipped", "not applicable"):
                status = "na"

        # Extract phase from column or detect from item text
        phase_col = col_map.get("phase")
        phase = ""
        if phase_col is not None and phase_col < len(cols):
            phase = cols[phase_col].strip().lower().replace(" ", "_")
        if not phase:
            phase = _detect_phase_from_text(item_text)

        # Extract priority from column or detect from item text
        priority_col = col_map.get("priority")
        priority = ""
        if priority_col is not None and priority_col < len(cols):
            priority = cols[priority_col].strip().lower()
        if not priority:
            priority = _extract_priority(item_text)

        # Extract assignee from column or detect from item text
        assignee_col = col_map.get("assignee")
        assignee = ""
        if assignee_col is not None and assignee_col < len(cols):
            assignee = cols[assignee_col].strip()
        if not assignee:
            assignee = _extract_assignee(item_text)

        items.append({
            "item_text": item_text,
            "status": status,
            "phase": phase,
            "priority": priority,
            "assignee": assignee,
            "section_heading": "",
            "context": "",
            "item_index": row_idx,
        })

    return items


def _parse_excel_checklist(binary: bytes) -> list[dict]:
    """Parse checklist items from Excel workbook.

    Expects columns: item/task, status, phase, priority, assignee.
    Handles merged cells and multi-sheet workbooks.

    Args:
        binary: Excel file content as bytes.

    Returns:
        List of dicts with checklist item metadata.
    """
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(binary), data_only=True)
    items = []
    item_index = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Detect header row
        header_keywords = {"item", "task", "checklist", "description", "status", "phase", "priority", "assignee"}
        first_row = [str(c).strip().lower() if c else "" for c in rows[0]]
        has_header = any(kw in cell for cell in first_row for kw in header_keywords)

        col_map = {"item": 0, "status": 1, "phase": 2, "priority": 3, "assignee": 4}
        start_row = 0

        if has_header:
            col_map = {}
            for idx, h in enumerate(first_row):
                if any(kw in h for kw in ("item", "task", "checklist", "description", "check")):
                    col_map["item"] = idx
                elif "status" in h or "done" in h or "complete" in h:
                    col_map["status"] = idx
                elif "phase" in h or "stage" in h:
                    col_map["phase"] = idx
                elif "priority" in h or "severity" in h:
                    col_map["priority"] = idx
                elif "assign" in h or "owner" in h or "responsible" in h:
                    col_map["assignee"] = idx
            if "item" not in col_map:
                col_map["item"] = 0
            start_row = 1

        for row in rows[start_row:]:
            cells = [str(c).strip() if c else "" for c in row]
            item_col = col_map.get("item", 0)
            item_text = cells[item_col] if item_col < len(cells) else ""
            if not item_text:
                continue

            # Parse status
            status_col = col_map.get("status")
            status = "unchecked"
            if status_col is not None and status_col < len(cells):
                raw = cells[status_col].lower()
                if raw in ("done", "yes", "x", "true", "checked", "pass", "passed", "complete", "completed"):
                    status = "checked"
                elif raw in ("n/a", "na", "skip", "skipped", "not applicable"):
                    status = "na"

            # Parse phase
            phase_col = col_map.get("phase")
            phase = ""
            if phase_col is not None and phase_col < len(cells):
                phase = cells[phase_col].lower().replace(" ", "_")
            if not phase:
                phase = _detect_phase_from_text(item_text)

            # Parse priority
            priority_col = col_map.get("priority")
            priority = ""
            if priority_col is not None and priority_col < len(cells):
                priority = cells[priority_col].lower()
            if not priority:
                priority = _extract_priority(item_text)

            # Parse assignee
            assignee_col = col_map.get("assignee")
            assignee = ""
            if assignee_col is not None and assignee_col < len(cells):
                assignee = cells[assignee_col]
            if not assignee:
                assignee = _extract_assignee(item_text)

            items.append({
                "item_text": item_text,
                "status": status,
                "phase": phase,
                "priority": priority,
                "assignee": assignee,
                "section_heading": sheet_name if len(wb.sheetnames) > 1 else "",
                "context": "",
                "item_index": item_index,
            })
            item_index += 1

    return items


def _parse_plaintext_checklist(text: str) -> list[dict]:
    """Parse checklist from plain text or PDF-extracted text.

    Detects multiple checklist formats:
    1. Markdown-style checkboxes: - [ ] item, - [x] item
    2. Bullet/numbered items: 1. item, - item, * item
    3. Status prefix patterns: [DONE] item, [PASS] item, [FAIL] item

    Falls back to paragraph-based splitting if no checklist patterns found.

    Args:
        text: Plain text content.

    Returns:
        List of dicts with checklist item metadata.
    """
    # Try Markdown checkbox format first
    if MD_CHECKBOX_PATTERN.search(text):
        return _parse_markdown_checklists(text)

    # Try status-prefix pattern: [DONE], [PASS], [FAIL], [TODO], [N/A]
    status_prefix_pattern = re.compile(
        r"^[\s]*(?:[-*+]|\d+[.)]\s*)?\s*\[(DONE|PASS|FAIL|TODO|N/?A|SKIP|PENDING|OK|NOK)\]\s*(.+)$",
        re.MULTILINE | re.IGNORECASE,
    )

    status_matches = status_prefix_pattern.findall(text)
    if status_matches:
        items = []
        # Detect overall phase from full text
        overall_phase = _detect_phase_from_text(text[:2000])

        sections = _split_md_sections(text)
        current_heading = ""
        current_phase = overall_phase

        for heading, content in sections:
            if heading:
                current_heading = heading
                detected = _detect_phase_from_heading(heading)
                if detected:
                    current_phase = detected

            for match in status_prefix_pattern.finditer(content):
                raw_status = match.group(1).upper()
                item_text = match.group(2).strip()

                # Map status prefix to normalized status
                if raw_status in ("DONE", "PASS", "OK"):
                    status = "checked"
                elif raw_status in ("N/A", "NA", "SKIP"):
                    status = "na"
                else:
                    status = "unchecked"

                items.append({
                    "item_text": item_text,
                    "status": status,
                    "phase": current_phase,
                    "priority": _extract_priority(item_text),
                    "assignee": _extract_assignee(item_text),
                    "section_heading": current_heading,
                    "context": "",
                    "item_index": len(items),
                })

        return items

    # Fallback: treat numbered/bullet items as unchecked checklist items
    bullet_pattern = re.compile(
        r"^[\s]*(?:[-*+]|\d+[.)]\s+)\s*(.+)$",
        re.MULTILINE,
    )
    bullet_matches = bullet_pattern.findall(text)
    if len(bullet_matches) >= 3:
        overall_phase = _detect_phase_from_text(text[:2000])
        items = []
        for idx, item_text in enumerate(bullet_matches):
            item_text = item_text.strip()
            if not item_text:
                continue
            items.append({
                "item_text": item_text,
                "status": "unchecked",
                "phase": _detect_phase_from_text(item_text) or overall_phase,
                "priority": _extract_priority(item_text),
                "assignee": _extract_assignee(item_text),
                "section_heading": "",
                "context": "",
                "item_index": idx,
            })
        return items

    return []


def _items_to_chunks(items: list[dict], doc: dict, group_by_section: bool = False) -> list[dict]:
    """Convert parsed checklist items into RAG chunk dicts.

    Each checklist item becomes a separate chunk with structured metadata
    fields for phase-based filtering and status tracking.

    Args:
        items: List of parsed checklist item dicts.
        doc: Base document metadata dict (docnm_kwd, title_tks).
        group_by_section: If True, group items by section heading into
            combined chunks instead of one-item-per-chunk.

    Returns:
        List of chunk dicts ready for indexing.
    """
    if not items:
        return []

    if group_by_section:
        return _group_items_by_section(items, doc)

    chunks = []
    for item in items:
        d = deepcopy(doc)

        # Build content with section context
        content_parts = []
        if item["section_heading"]:
            content_parts.append(f"[{item['section_heading']}]")
        content_parts.append(item["item_text"])
        if item.get("context"):
            content_parts.append(item["context"])

        chunk_text = "\n".join(content_parts)

        d["content_with_weight"] = chunk_text
        d["content_ltks"] = rag_tokenizer.tokenize(chunk_text)
        d["content_sm_ltks"] = rag_tokenizer.fine_grained_tokenize(d["content_ltks"])

        # Structured metadata fields
        d["tag_kwd"] = ["sdlc_checklist"]
        if item["phase"]:
            d["tag_kwd"].append(item["phase"])
        d["sdlc_phase_kwd"] = item["phase"] if item["phase"] else "unclassified"
        d["sdlc_status_kwd"] = item["status"]
        if item["priority"]:
            d["sdlc_priority_kwd"] = item["priority"]
        if item["assignee"]:
            d["sdlc_assignee_kwd"] = item["assignee"]
        d["top_int"] = [item["item_index"]]

        chunks.append(d)

    return chunks


def _group_items_by_section(items: list[dict], doc: dict) -> list[dict]:
    """Group checklist items by section heading into combined chunks.

    Used when individual items are too small to be useful as standalone
    chunks. Groups items under their section heading and computes
    aggregate completion stats.

    Args:
        items: List of parsed checklist item dicts.
        doc: Base document metadata dict.

    Returns:
        List of section-level chunk dicts.
    """
    sections: dict[str, list[dict]] = {}
    for item in items:
        key = item.get("section_heading", "") or "General"
        sections.setdefault(key, []).append(item)

    chunks = []
    for section_name, section_items in sections.items():
        d = deepcopy(doc)

        # Build section content with all items
        lines = [f"## {section_name}"]
        checked_count = 0
        total_count = len(section_items)

        for item in section_items:
            marker = "x" if item["status"] == "checked" else ("~" if item["status"] == "na" else " ")
            lines.append(f"- [{marker}] {item['item_text']}")
            if item["status"] == "checked":
                checked_count += 1

        chunk_text = "\n".join(lines)

        d["content_with_weight"] = chunk_text
        d["content_ltks"] = rag_tokenizer.tokenize(chunk_text)
        d["content_sm_ltks"] = rag_tokenizer.fine_grained_tokenize(d["content_ltks"])

        # Detect dominant phase for the section
        phase = section_items[0]["phase"] if section_items else ""
        d["tag_kwd"] = ["sdlc_checklist"]
        if phase:
            d["tag_kwd"].append(phase)
        d["sdlc_phase_kwd"] = phase if phase else "unclassified"

        # Aggregate completion stats
        d["sdlc_checked_int"] = checked_count
        d["sdlc_total_int"] = total_count
        d["top_int"] = [section_items[0]["item_index"]] if section_items else [0]

        chunks.append(d)

    return chunks


def chunk(filename, binary=None, from_page=0, to_page=100000,
          lang="Chinese", callback=None, **kwargs):
    """Parse an SDLC checklist document into per-item chunks.

    Supports multiple formats: Markdown (.md), Excel (.xlsx), CSV (.csv),
    TXT (.txt), PDF (.pdf), and DOCX (.docx). Detects checklist patterns
    (checkboxes, status prefixes, tabular layouts) and extracts structured
    metadata per item including SDLC phase, completion status, priority,
    and assignee.

    Args:
        filename: Original filename (used for format detection).
        binary: Raw file content as bytes.
        from_page: Starting page for PDF parsing.
        to_page: Ending page for PDF parsing.
        lang: Language hint for tokenization ('English', 'Chinese', etc.).
        callback: Progress callback function(progress_float, message_str).
        **kwargs: Additional config including parser_config with:
            - chunk_token_num: Target tokens per chunk (default 512).
            - group_by_section: Group items by section heading (default False).
            - sdlc_phase: Override phase detection with explicit phase.

    Returns:
        List of chunk dicts, each containing content_with_weight, tokenized
        fields, and SDLC metadata (sdlc_phase_kwd, sdlc_status_kwd,
        sdlc_priority_kwd, sdlc_assignee_kwd, tag_kwd).
    """
    parser_config = kwargs.get("parser_config", {})
    group_by_section = parser_config.get("group_by_section", False)
    explicit_phase = parser_config.get("sdlc_phase", "")

    if callback is None:
        callback = lambda prog, msg="": None

    # Handle empty input
    if not binary:
        return []

    # Base document metadata
    doc = {
        "docnm_kwd": filename,
        "title_tks": rag_tokenizer.tokenize(re.sub(r"\.[a-zA-Z]+$", "", filename)),
    }

    callback(0.1, "Start to parse SDLC checklist document.")

    items = []

    # Route to format-specific parser
    if re.search(r"\.xlsx?$", filename, re.IGNORECASE):
        items = _parse_excel_checklist(binary)
        callback(0.5, f"Extracted {len(items)} checklist items from Excel.")

    elif re.search(r"\.(csv)$", filename, re.IGNORECASE):
        text = get_text(filename, binary)
        # Auto-detect delimiter
        if not text:
            callback(1.0, "Empty document.")
            return []
        first_lines = text.split("\n")[:5]
        tab_count = sum(line.count("\t") for line in first_lines)
        comma_count = sum(line.count(",") for line in first_lines)
        delimiter = "\t" if tab_count >= comma_count else ","
        items = _parse_tabular_checklist(text, delimiter)
        callback(0.5, f"Extracted {len(items)} checklist items from CSV.")

    elif re.search(r"\.(md|markdown|mdx)$", filename, re.IGNORECASE):
        try:
            text = binary.decode("utf-8")
        except UnicodeDecodeError:
            text = binary.decode("utf-8", errors="replace")
        items = _parse_markdown_checklists(text)
        callback(0.5, f"Extracted {len(items)} checklist items from Markdown.")

    elif re.search(r"\.(txt)$", filename, re.IGNORECASE):
        text = get_text(filename, binary)
        if not text:
            callback(1.0, "Empty document.")
            return []
        # Try tabular first (TSV/CSV), then plaintext patterns
        tab_count = text.count("\t")
        if tab_count > 5:
            items = _parse_tabular_checklist(text, "\t")
        else:
            items = _parse_plaintext_checklist(text)
        callback(0.5, f"Extracted {len(items)} checklist items from text.")

    elif re.search(r"\.pdf$", filename, re.IGNORECASE):
        text = get_text(filename, binary)
        if not text:
            callback(1.0, "Empty document.")
            return []
        items = _parse_plaintext_checklist(text)
        callback(0.5, f"Extracted {len(items)} checklist items from PDF.")

    elif re.search(r"\.docx?$", filename, re.IGNORECASE):
        text = get_text(filename, binary)
        if not text:
            callback(1.0, "Empty document.")
            return []
        items = _parse_plaintext_checklist(text)
        callback(0.5, f"Extracted {len(items)} checklist items from DOCX.")

    else:
        # Unsupported format — try plain text as last resort
        text = get_text(filename, binary)
        if text:
            items = _parse_plaintext_checklist(text)

    # Apply explicit phase override if provided
    if explicit_phase:
        for item in items:
            item["phase"] = explicit_phase

    # Fallback: if no structured items found, use paragraph-based chunking
    if not items:
        callback(0.6, "No checklist pattern detected, falling back to paragraph chunking.")
        return _fallback_paragraph_chunking(filename, binary, doc, parser_config, callback)

    # Convert items to chunks
    res = _items_to_chunks(items, doc, group_by_section)

    # Compute document-level completion stats
    total = len(items)
    checked = sum(1 for i in items if i["status"] == "checked")
    na_count = sum(1 for i in items if i["status"] == "na")
    effective_total = total - na_count

    completion_pct = (checked / effective_total * 100) if effective_total > 0 else 0.0

    callback(0.9, (
        f"SDLC checklist parsed: {len(res)} chunks, "
        f"{checked}/{effective_total} items checked ({completion_pct:.0f}% complete)."
    ))

    return res


def _fallback_paragraph_chunking(filename: str, binary: bytes, doc: dict,
                                 parser_config: dict, callback) -> list[dict]:
    """Fall back to paragraph-based chunking when no checklist pattern found.

    Uses the same approach as the clinical parser: split by blank lines,
    merge paragraphs up to token limit.

    Args:
        filename: Original filename.
        binary: Raw file content as bytes.
        doc: Base document metadata.
        parser_config: Parser configuration dict.
        callback: Progress callback function.

    Returns:
        List of chunk dicts with sdlc_checklist tag.
    """
    chunk_token_num = int(parser_config.get("chunk_token_num", DEFAULT_CHUNK_TOKEN_NUM))

    text = get_text(filename, binary)
    if not text or not text.strip():
        callback(1.0, "Empty document.")
        return []

    # Split into paragraphs by blank lines
    paragraphs = [p.strip() for p in re.split(r"\n\s*\n", text) if p.strip()]

    # Detect overall phase from full text
    overall_phase = _detect_phase_from_text(text[:2000])

    # Merge paragraphs into chunks respecting token limit
    chunks_text = []
    current = ""
    current_tokens = 0

    for para in paragraphs:
        para_tokens = num_tokens_from_string(para)
        if current and (current_tokens + para_tokens) > chunk_token_num:
            chunks_text.append(current.strip())
            current = para
            current_tokens = para_tokens
        else:
            current = (current + "\n\n" + para) if current else para
            current_tokens += para_tokens

    if current.strip():
        chunks_text.append(current.strip())

    # Build chunk dicts
    res = []
    for i, chunk_text in enumerate(chunks_text):
        d = deepcopy(doc)
        d["content_with_weight"] = chunk_text
        d["content_ltks"] = rag_tokenizer.tokenize(chunk_text)
        d["content_sm_ltks"] = rag_tokenizer.fine_grained_tokenize(d["content_ltks"])
        d["tag_kwd"] = ["sdlc_checklist"]
        if overall_phase:
            d["tag_kwd"].append(overall_phase)
        d["sdlc_phase_kwd"] = overall_phase if overall_phase else "unclassified"
        d["top_int"] = [i]
        res.append(d)

    callback(0.9, f"Fallback paragraph chunking: {len(res)} chunks.")
    return res


async def classify_phase(chat_mdl, full_text: str, title: str) -> str:
    """Classify an SDLC checklist document into an SDLC phase using LLM.

    Called from task_executor.py as a post-parse step when keyword-based
    phase detection produces 'unclassified'. Sends the title and first
    2000 tokens to the LLM. Falls back to 'unclassified' on failure.

    Args:
        chat_mdl: LLMBundle instance for chat completion.
        full_text: Full document text content.
        title: Document title or filename.

    Returns:
        Phase identifier string (one of SDLC_PHASES keys or 'unclassified').
    """
    valid_phases = set(SDLC_PHASES.keys())

    try:
        # Truncate content to first MAX_CLASSIFY_TOKENS tokens
        words = full_text.split()
        if len(words) > MAX_CLASSIFY_TOKENS:
            content_preview = " ".join(words[:MAX_CLASSIFY_TOKENS])
        else:
            content_preview = full_text

        prompt = PHASE_CLASSIFICATION_PROMPT.format(
            title=title,
            content=content_preview,
        )

        result = await chat_mdl.async_chat(prompt, [{"role": "user", "content": prompt}], {})

        # Handle tuple response from some LLM implementations
        if isinstance(result, tuple):
            result = result[0] if result else ""

        phase = result.strip().lower().replace(" ", "_") if result else ""

        if phase in valid_phases:
            return phase

        logger.warning(
            "SDLC phase classifier returned invalid phase '%s' for '%s', defaulting to 'unclassified'",
            phase, title,
        )
        return "unclassified"

    except Exception as e:
        logger.warning(
            "SDLC phase classification failed for '%s': %s. Defaulting to 'unclassified'.",
            title, str(e),
        )
        return "unclassified"
